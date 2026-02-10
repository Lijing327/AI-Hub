"""
Excel 导入业务：读取 Excel、行映射、调用 .NET 批量创建文章与附件
"""
from io import BytesIO
from typing import List, Optional

import pandas as pd
import httpx
from fastapi import HTTPException

from app.core.config import settings
from app.core.logging_config import get_logger
from app.clients.dotnet_client import DotnetClient
from app.schemas.excel import ExcelImportResponse
from app.services.attachment_service import AttachmentService
from app.services.excel_utils import map_excel_row_to_article

logger = get_logger(__name__)


class ExcelImportService:
    """Excel 导入服务：解析表头与行，调用 .NET 批量创建并顺带创建附件"""

    def __init__(
        self,
        dotnet_client: Optional[DotnetClient] = None,
        attachment_service: Optional[AttachmentService] = None,
    ):
        self.dotnet_client = dotnet_client or DotnetClient()
        self.attachment_service = attachment_service or AttachmentService()

    async def import_excel(self, file_content: bytes, filename: str) -> ExcelImportResponse:
        """
        导入 Excel 为知识条目
        - 校验 .xlsx、解析表头与行
        - 每行映射为一条知识草稿（含 _attachment_info）
        - 调用 .NET 批量创建文章，再批量创建附件
        """
        if not filename or not filename.endswith(".xlsx"):
            raise HTTPException(status_code=400, detail="只支持 .xlsx 格式的 Excel 文件")

        excel_file = BytesIO(file_content)
        df = pd.read_excel(excel_file, engine="openpyxl", header=0)

        fault_phenomenon_columns = [
            "故障现象", "现象（问题）", "现象(问题)", "现象 （问题）", "现象 (问题)", "现象", "问题",
        ]

        def normalize_col(c: str) -> str:
            return str(c).strip().lower()

        norm_columns = {normalize_col(c): c for c in df.columns}
        norm_fault = [normalize_col(c) for c in fault_phenomenon_columns]
        has_fault = any(n in norm_columns for n in norm_fault)
        header_row = 0

        if not has_fault:
            excel_file.seek(0)
            df = pd.read_excel(excel_file, engine="openpyxl", header=1)
            header_row = 1
            norm_columns = {normalize_col(c): c for c in df.columns}
            has_fault = any(n in norm_columns for n in norm_fault)
            if not has_fault:
                for col in df.columns:
                    cl = str(col).strip().lower()
                    if "现象" in cl or "问题" in cl or "故障" in cl:
                        has_fault = True
                        break
            if not has_fault:
                raise HTTPException(
                    status_code=400,
                    detail=f"缺少必需字段，请包含以下任一列：{', '.join(fault_phenomenon_columns)}；实际列名：{list(df.columns)}",
                )

        if df.empty:
            raise HTTPException(status_code=400, detail="Excel 文件为空")

        # 确定“现象”列名（用于后面过滤无效行）
        phenomenon_col = None
        for v in fault_phenomenon_columns:
            nv = normalize_col(v)
            if nv in norm_columns:
                phenomenon_col = norm_columns[nv]
                break
        if not phenomenon_col and has_fault:
            for c in df.columns:
                if "现象" in str(c) or "问题" in str(c) or "故障" in str(c):
                    phenomenon_col = c
                    break

        # 使用 header=1 时，首行多为标题行，丢弃避免多导 1 条
        if header_row == 1 and len(df) > 0:
            df = df.iloc[1:].reset_index(drop=True)
            if df.empty:
                raise HTTPException(status_code=400, detail="Excel 除标题行外无数据")

        # 只保留“现象”列有实质内容的行，避免空行、重复表头行被当作数据（导致 45 条变 50 条等）
        header_like = {"序号", "现象", "问题", "检查点", "原因", "维修对策", "解决办法", "维修视频", "附件", "故障现象"}
        if phenomenon_col and phenomenon_col in df.columns:
            def _valid_phenomenon(val) -> bool:
                if pd.isna(val):
                    return False
                s = str(val).strip()
                if not s or s.lower() == "nan":
                    return False
                if s in header_like or len(s) < 2:
                    return False
                return True
            df = df[df[phenomenon_col].apply(_valid_phenomenon)].reset_index(drop=True)

        if df.empty:
            raise HTTPException(status_code=400, detail="没有有效数据行（现象列为空或仅为表头）")

        # 过滤“型号”列
        cols_drop = [c for c in df.columns if "型号" in str(c).strip()]
        if cols_drop:
            df = df.drop(columns=cols_drop)

        sheet_name = "Sheet1"
        try:
            import openpyxl
            excel_file.seek(0)
            wb = openpyxl.load_workbook(excel_file, read_only=True)
            if wb.sheetnames:
                sheet_name = wb.sheetnames[0]
            excel_file.seek(0)
        except Exception as e:
            logger.warning("无法获取 sheet 名称，使用默认: %s", e)

        articles: List[dict] = []
        articles_with_attachments: List[dict] = []
        failures: List[dict] = []
        skipped = 0

        for idx, row in df.iterrows():
            excel_row_num = int(idx) + header_row + 2
            try:
                article = map_excel_row_to_article(
                    row, filename, sheet_name, excel_row_num, self.attachment_service,
                )
                if article:
                    articles_with_attachments.append(article)
                    api_article = {k: v for k, v in article.items() if not str(k).startswith("_")}
                    articles.append(api_article)
                else:
                    skipped += 1
            except Exception as e:
                logger.exception("处理第 %s 行出错", excel_row_num)
                failures.append({"row_index": excel_row_num, "reason": str(e)})

        if not articles:
            raise HTTPException(
                status_code=400,
                detail=f"没有有效数据行。总行数: {len(df)}, 跳过: {skipped}, 失败: {len(failures)}；列名: {list(df.columns)}",
            )

        logger.info("准备调用 .NET 批量创建文章，数量: %s", len(articles))
        try:
            result = await self.dotnet_client.batch_create_articles(articles)
        except httpx.ConnectError as e:
            logger.error("无法连接 .NET 后端: %s", e)
            raise HTTPException(status_code=503, detail=f"无法连接 .NET 后端: {settings.DOTNET_BASE_URL}")
        except httpx.TimeoutException as e:
            logger.error("调用 .NET 超时: %s", e)
            raise HTTPException(status_code=504, detail="调用 .NET 后端超时，请稍后重试")
        except httpx.HTTPStatusError as e:
            logger.error(".NET 返回错误: %s %s", e.response.status_code, e.response.text)
            raise HTTPException(
                status_code=e.response.status_code,
                detail=f".NET 后端错误: {e.response.text}",
            )

        success_count = result.get("successCount", 0)
        failure_count = result.get("failureCount", 0)
        article_ids: List[int] = []
        article_id_to_attachments: dict = {}

        for i, item in enumerate(result.get("results", [])):
            if item.get("success") and item.get("articleId"):
                aid = item["articleId"]
                article_ids.append(aid)
                if i < len(articles_with_attachments):
                    att = articles_with_attachments[i].get("_attachment_info")
                    if att:
                        article_id_to_attachments[aid] = att if isinstance(att, list) else [att]
            elif not item.get("success"):
                failures.append({
                    "row_index": item.get("index", -1) + header_row + 2,
                    "reason": item.get("error", "未知错误"),
                })

        # 批量创建附件
        if article_id_to_attachments:
            assets_to_create: List[dict] = []
            for aid, att_list in article_id_to_attachments.items():
                for info in att_list:
                    assets_to_create.append({
                        "articleId": aid,
                        "assetType": info.get("asset_type", "other"),
                        "fileName": info.get("file_name", ""),
                        "url": info.get("url", ""),
                        "size": info.get("size"),
                        "duration": None,
                    })
            if assets_to_create:
                try:
                    asset_result = await self.dotnet_client.batch_create_assets(assets_to_create)
                    logger.info(
                        "附件创建结果: 成功 %s, 失败 %s",
                        asset_result.get("successCount", 0),
                        asset_result.get("failureCount", 0),
                    )
                except Exception as e:
                    logger.exception("批量创建附件失败: %s", e)

        return ExcelImportResponse(
            total_rows=len(df),
            success_count=success_count,
            failure_count=len(failures) + failure_count,
            article_ids=article_ids,
            failures=failures,
        )
