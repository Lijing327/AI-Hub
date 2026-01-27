"""
AI Hub Excel 导入服务
FastAPI 服务，用于处理 Excel 导入并调用 .NET 后端 API
"""
import os
import traceback
from typing import List, Optional
from io import BytesIO
from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.responses import JSONResponse
from pydantic import BaseModel
import pandas as pd
import httpx
from pathlib import Path
import json
import re
import logging
from urllib.parse import quote

# 配置日志
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = FastAPI(
    title="AI Hub Excel 导入服务",
    description="处理 Excel 文件导入，转换为知识条目",
    version="1.0.0"
)

# 配置（从环境变量读取）
DOTNET_BASE_URL = os.getenv("DOTNET_BASE_URL", "http://localhost:5000")
INTERNAL_TOKEN = os.getenv("INTERNAL_TOKEN", "your-internal-token-change-in-production")
DEFAULT_TENANT = os.getenv("DEFAULT_TENANT", "default")
ATTACHMENT_BASE_PATH = os.getenv("ATTACHMENT_BASE_PATH", "")
ATTACHMENT_BASE_URL = os.getenv("ATTACHMENT_BASE_URL", "http://localhost:5000/uploads")

# 加载 .env 文件（如果存在）
try:
    from dotenv import load_dotenv
    load_dotenv()
    # 重新读取环境变量（.env 文件中的值会覆盖上面的默认值）
    DOTNET_BASE_URL = os.getenv("DOTNET_BASE_URL", DOTNET_BASE_URL)
    INTERNAL_TOKEN = os.getenv("INTERNAL_TOKEN", INTERNAL_TOKEN)
    DEFAULT_TENANT = os.getenv("DEFAULT_TENANT", DEFAULT_TENANT)
    ATTACHMENT_BASE_PATH = os.getenv("ATTACHMENT_BASE_PATH", ATTACHMENT_BASE_PATH)
    ATTACHMENT_BASE_URL = os.getenv("ATTACHMENT_BASE_URL", ATTACHMENT_BASE_URL)
    logger.info(f"配置加载完成: DOTNET_BASE_URL={DOTNET_BASE_URL}, DEFAULT_TENANT={DEFAULT_TENANT}, ATTACHMENT_BASE_PATH={ATTACHMENT_BASE_PATH}")
except ImportError:
    logger.warning("python-dotenv 未安装，将使用环境变量或默认值")


class ExcelImportResponse(BaseModel):
    """Excel 导入响应"""
    total_rows: int
    success_count: int
    failure_count: int
    article_ids: List[int]
    failures: List[dict]


class ExcelRowFailure(BaseModel):
    """Excel 行处理失败信息"""
    row_index: int
    reason: str


def format_as_reasons(text: Optional[str]) -> Optional[str]:
    """格式化为原因列表（原因 1：...）"""
    if not text or not text.strip():
        return None
    
    # 按换行符或分号分割
    reasons = [r.strip() for r in re.split(r'[\n\r；;]', text) if r.strip()]
    
    if not reasons:
        return None
    
    # 如果已经是"原因 X："格式，直接返回
    if any(re.match(r'^原因\s*\d+[：:]', r) for r in reasons):
        return '\n'.join(reasons)
    
    # 否则添加编号
    return '\n'.join([f"原因 {i + 1}：{r}" for i, r in enumerate(reasons)])


def format_as_steps(text: Optional[str]) -> Optional[str]:
    """格式化为步骤列表（步骤 1：...）"""
    if not text or not text.strip():
        return None
    
    # 按换行符或分号分割
    steps = [s.strip() for s in re.split(r'[\n\r；;]', text) if s.strip()]
    
    if not steps:
        return None
    
    # 如果已经是"步骤 X："格式，直接返回
    if any(re.match(r'^步骤\s*\d+[：:]', s) for s in steps):
        return '\n'.join(steps)
    
    # 否则添加编号
    return '\n'.join([f"步骤 {i + 1}：{s}" for i, s in enumerate(steps)])


def clean_bracketed_labels(text: str) -> str:
    """
    清理文本中的中括号标签和列标题
    移除：【序号】、【现象（问题）】、【检查点（原因）】、【维修对策（解决办法）】、【维修视频（附件）】、【YH400/YH500】等
    
    支持多种格式：
    - 【序号】、【现象（问题）】等（中文括号）
    - [序号]、[现象(问题)]等（英文括号）
    - 各种变体：现象(问题)、现象（问题）、现象 (问题)等
    """
    if not text or not isinstance(text, str):
        return text if text else ""
    
    import re
    
    # 更全面的模式匹配，支持各种变体
    # 使用更简单直接的方式：匹配所有包含这些关键词的中括号内容
    patterns_to_remove = [
        # 匹配任何包含"序号"的中括号
        r'【[^】]*序号[^】]*】',
        # 匹配任何包含"现象"和"问题"的中括号（支持各种变体）
        r'【[^】]*现象[^】]*问题[^】]*】',
        r'【[^】]*现象[^】]*】',
        r'【[^】]*问题[^】]*】',
        # 匹配任何包含"检查点"和"原因"的中括号
        r'【[^】]*检查点[^】]*原因[^】]*】',
        r'【[^】]*检查点[^】]*】',
        r'【[^】]*原因[^】]*】',
        # 匹配任何包含"维修对策"和"解决办法"的中括号
        r'【[^】]*维修对策[^】]*解决办法[^】]*】',
        r'【[^】]*维修对策[^】]*】',
        r'【[^】]*解决办法[^】]*】',
        # 匹配任何包含"维修视频"和"附件"的中括号
        r'【[^】]*维修视频[^】]*附件[^】]*】',
        r'【[^】]*维修视频[^】]*】',
        r'【[^】]*附件[^】]*】',
        # 匹配型号信息
        r'【[^】]*YH400[^】]*YH500[^】]*】',
        r'【[^】]*YH400[^】]*】',
        r'【[^】]*YH500[^】]*】',
        # 英文括号格式（同样处理）
        r'\[[^\]]*序号[^\]]*\]',
        r'\[[^\]]*现象[^\]]*问题[^\]]*\]',
        r'\[[^\]]*检查点[^\]]*原因[^\]]*\]',
        r'\[[^\]]*维修对策[^\]]*解决办法[^\]]*\]',
        r'\[[^\]]*维修视频[^\]]*附件[^\]]*\]',
        r'\[[^\]]*YH400[^\]]*YH500[^\]]*\]',
        r'\[[^\]]*YH400[^\]]*\]',
        r'\[[^\]]*YH500[^\]]*\]',
    ]
    
    cleaned_text = text
    # 多次清理，确保所有匹配都被移除
    for _ in range(5):  # 循环5次，确保嵌套或重复的标签都被清理
        old_text = cleaned_text
        for pattern in patterns_to_remove:
            cleaned_text = re.sub(pattern, '', cleaned_text, flags=re.IGNORECASE)
        # 如果这一轮没有变化，提前退出
        if cleaned_text == old_text:
            break
    
    # 清理多余的空白字符（但保留换行）
    # 先清理连续的空格
    cleaned_text = re.sub(r'[ \t]+', ' ', cleaned_text)
    # 清理行首行尾空格（但保留换行）
    lines = cleaned_text.split('\n')
    cleaned_lines = [line.strip() for line in lines]
    cleaned_text = '\n'.join(cleaned_lines)
    # 清理多余的连续换行（最多保留一个空行）
    cleaned_text = re.sub(r'\n{3,}', '\n\n', cleaned_text)
    cleaned_text = cleaned_text.strip()
    
    return cleaned_text


def extract_filename_from_reference(text: str) -> Optional[str]:
    """
    从文本引用中提取文件名
    支持格式：
    - 参考"xxx" 或 参考"xxx"（中文引号 "" 和英文引号 ""）
    - 参考：xxx
    - 见附件：xxx
    - 参考 xxx
    """
    if not text or not text.strip():
        return None
    
    text = text.strip()
    
    # 匹配：参考"xxx" 或 参考"xxx"（支持中文引号 "" 和英文引号 ""）
    # 正则：参考[""]... [""]（支持中文引号和英文引号）
    # 使用字符类匹配所有类型的引号
    match = re.search(r'参考["""""]([^"""""]+)["""""]', text)
    if match:
        filename = match.group(1).strip()
        # 去除可能残留的引号（包括全角单引号 ''）
        filename = filename.strip('"""\'""\'')
        return filename if filename else None
    
    # 匹配：参考：xxx 或 参考:xxx
    match = re.search(r'参考[：:]\s*(.+)', text)
    if match:
        filename = match.group(1).strip()
        # 去除可能的前后引号（包括全角单引号 ''）
        filename = filename.strip('"""\'""\'')
        return filename if filename else None
    
    # 匹配：见附件：xxx
    match = re.search(r'见附件[：:]\s*(.+)', text)
    if match:
        filename = match.group(1).strip()
        filename = filename.strip('"""\'""\'')
        return filename if filename else None
    
    # 匹配：参考 xxx（空格分隔）
    match = re.search(r'参考\s+(.+)', text)
    if match:
        filename = match.group(1).strip()
        filename = filename.strip('"""\'""\'')
        return filename if filename else None
    
    # 如果都不匹配，返回原文（去除"参考"等前缀）
    filename = re.sub(r'^(参考|见附件)[：:\s]*', '', text)
    filename = filename.strip('"""\'""\'')
    return filename.strip() if filename.strip() else None


def _guess_asset_type_by_ext(ext: str) -> str:
    """根据扩展名判断文件类型"""
    ext = ext.lower()
    if ext in ['.mp4', '.avi', '.mov', '.wmv', '.flv', '.mkv']:
        return 'video'
    if ext in ['.jpg', '.jpeg', '.png', '.gif', '.bmp', '.webp']:
        return 'image'
    if ext in ['.pdf']:
        return 'pdf'
    if ext in ['.doc', '.docx', '.xls', '.xlsx', '.txt', '.ppt', '.pptx']:
        return 'other'
    return 'other'


def _should_skip_file(p: Path) -> bool:
    """过滤系统/临时文件，避免脏数据"""
    name = p.name.lower()
    if name in ['thumbs.db', '.ds_store', 'desktop.ini']:
        return True
    if name.startswith('~$'):  # office 临时文件
        return True
    return False


def _build_file_info(base_path: Path, file_path: Path) -> dict:
    """构建文件信息字典"""
    file_size = file_path.stat().st_size
    relative_path = file_path.relative_to(base_path)
    relative_path_str = str(relative_path).replace(os.sep, '/')
    encoded_path = '/'.join(quote(part, safe='') for part in relative_path_str.split('/'))
    file_url = f"{ATTACHMENT_BASE_URL.rstrip('/')}/{encoded_path}"
    return {
        "path": str(file_path),
        "url": file_url,
        "type": _guess_asset_type_by_ext(file_path.suffix),
        "size": file_size,
        "file_name": file_path.name,
        "relative_path": relative_path_str,
    }


def find_attachment_files(filename: str) -> List[dict]:
    """
    在固定目录中递归查找附件文件（支持文件夹嵌套）
    返回：List[{ "path","url","type","size","file_name", "relative_path" }]
    - 命中文件：返回 1 条
    - 命中文件夹：返回该文件夹内所有文件（递归）的多条
    - 未命中：返回空列表
    """
    if not ATTACHMENT_BASE_PATH or not ATTACHMENT_BASE_PATH.strip():
        logger.debug(f"ATTACHMENT_BASE_PATH 未配置，跳过文件查找: {filename}")
        return []

    base_path = Path(ATTACHMENT_BASE_PATH.strip())
    if not base_path.exists():
        logger.debug(f"附件基础路径不存在: {base_path}，跳过文件查找: {filename}")
        return []

    # 清理文件名：去除引号、括号等包裹符号（包括全角单引号 ''）
    clean_filename = filename.strip().strip('"""\'""\'《》【】[]()（）').strip()
    logger.debug(f"开始查找附件: 原始文件名='{filename}', 清理后='{clean_filename}'")

    # 如果用户写了扩展名，先转 stem（Excel 里可能不带扩展名）
    if '.' in clean_filename:
        original_clean = clean_filename
        clean_filename = Path(clean_filename).stem
        logger.debug(f"检测到扩展名，提取stem: '{original_clean}' -> '{clean_filename}'")

    # 扩展名映射（用于"文件命中"阶段）
    extensions_map = {
        'video': ['.mp4', '.avi', '.mov', '.wmv', '.flv', '.mkv'],
        'image': ['.jpg', '.jpeg', '.png', '.gif', '.bmp', '.webp'],
        'pdf': ['.pdf'],
        'other': ['.doc', '.docx', '.xls', '.xlsx', '.txt', '.ppt', '.pptx']
    }

    # ========== 1) 精确匹配 / 递归精确匹配（命中文件就返回 1 条） ==========
    for asset_type, extensions in extensions_map.items():
        for ext in extensions:
            full_filename = clean_filename + ext
            logger.debug(f"尝试精确匹配: '{full_filename}' (类型: {asset_type}, 扩展名: {ext})")

            # 根目录精确匹配
            root_file = base_path / full_filename
            if root_file.is_file():
                logger.info(f"找到附件文件（根目录精确匹配）: {clean_filename} -> {root_file.name}")
                info = _build_file_info(base_path, root_file)
                info["type"] = asset_type  # 按映射强制类型
                return [info]
            else:
                logger.debug(f"  根目录精确匹配失败: {root_file} (存在: {root_file.exists()}, 是文件: {root_file.is_file() if root_file.exists() else False})")

            # 全目录精确匹配
            matches = list(base_path.rglob(full_filename))
            logger.debug(f"  全目录精确匹配: 找到 {len(matches)} 个匹配项")
            for file_path in matches:
                if file_path.is_file():
                    logger.info(f"找到附件文件（精确匹配）: {clean_filename} -> {file_path.name}")
                    info = _build_file_info(base_path, file_path)
                    info["type"] = asset_type
                    return [info]

            # ========== 2) 模糊匹配（命中文件就返回 1 条） ==========
            pattern = f"*{clean_filename}*{ext}"

            root_matches = list(base_path.glob(pattern))
            matches = root_matches + list(base_path.rglob(pattern))

            # 去重
            seen = set()
            unique_matches = []
            for m in matches:
                if m not in seen:
                    seen.add(m)
                    unique_matches.append(m)
            matches = unique_matches

            if matches:
                best_match = None
                clean_lower = clean_filename.lower()

                for p in matches:
                    if p.is_file() and clean_lower in p.stem.lower():
                        best_match = p
                        break

                if not best_match:
                    for p in matches:
                        if p.is_file():
                            best_match = p
                            break

                if best_match and best_match.is_file():
                    logger.info(f"找到附件文件（模糊匹配）: {clean_filename} -> {best_match.name}")
                    info = _build_file_info(base_path, best_match)
                    info["type"] = asset_type
                    return [info]

    # ========== 3) 文件夹命中：返回该文件夹内所有文件（递归，多条） ==========
    logger.debug(f"文件匹配失败，尝试匹配文件夹名称: {clean_filename}")
    clean_lower = clean_filename.lower()

    matched_folders = []
    for folder_path in base_path.rglob("*"):
        if not folder_path.is_dir():
            continue
        folder_lower = folder_path.name.lower()

        # 完全匹配优先
        if clean_lower == folder_lower:
            matched_folders.insert(0, (folder_path, 1))
        # 包含匹配兜底
        elif clean_lower in folder_lower or folder_lower in clean_lower:
            matched_folders.append((folder_path, 2))

    matched_folders.sort(key=lambda x: x[1])

    for folder_path, priority in matched_folders:
        # 递归收集文件夹内所有文件
        all_files: List[Path] = []
        for p in folder_path.rglob("*"):
            if p.is_file() and not _should_skip_file(p):
                all_files.append(p)

        if not all_files:
            continue

        # 排序保证稳定（方便比对导入结果）
        all_files.sort(key=lambda p: str(p).lower())

        results = []
        for p in all_files:
            info = _build_file_info(base_path, p)
            results.append(info)

        match_type = "完全匹配" if priority == 1 else "包含匹配"
        logger.info(
            f"找到附件文件（文件夹{match_type}）: {clean_filename} -> 文件夹[{folder_path.name}] 共 {len(results)} 个文件"
        )
        return results

    # P1: 未命中时打印"根目录/全目录"候选 stem（限制 5 个）
    logger.warning(f"未找到附件文件或文件夹: {filename} (清理后: {clean_filename}, 搜索路径: {base_path})")
    
    # 列出根目录下所有文件（用于调试）
    try:
        root_files = [f.name for f in base_path.iterdir() if f.is_file()]
        logger.debug(f"  根目录文件列表（前10个）: {root_files[:10]}")
        
        # 查找包含关键字的文件
        root_candidates = []
        for file_path in base_path.iterdir():
            if file_path.is_file():
                stem_lower = file_path.stem.lower()
                clean_lower = clean_filename.lower()
                if clean_lower in stem_lower or stem_lower in clean_lower:
                    root_candidates.append(f"{file_path.stem}{file_path.suffix}")
                    if len(root_candidates) >= 5:
                        break
        if root_candidates:
            logger.info(f"  根目录候选文件: {root_candidates}")
        else:
            logger.debug(f"  根目录未找到包含 '{clean_filename}' 的文件")
    except Exception as e:
        logger.debug(f"  无法列出根目录候选文件: {str(e)}")
    
    # 列出全目录下可能的候选文件（限制 5 个）
    try:
        all_candidates = []
        for file_path in base_path.rglob("*"):
            if file_path.is_file():
                stem_lower = file_path.stem.lower()
                clean_lower = clean_filename.lower()
                if clean_lower in stem_lower or stem_lower in clean_lower:
                    all_candidates.append(f"{file_path.stem}{file_path.suffix} (路径: {file_path.relative_to(base_path)})")
                    if len(all_candidates) >= 5:
                        break
        if all_candidates:
            logger.info(f"  全目录候选文件: {all_candidates}")
        else:
            logger.debug(f"  全目录未找到包含 '{clean_filename}' 的文件")
    except Exception as e:
        logger.debug(f"  无法列出全目录候选文件: {str(e)}")
    
    return []


def find_column_by_variants(row: pd.Series, variants: List[str]) -> Optional[str]:
    """通过多种变体查找列，支持忽略空格和大小写"""
    # 先尝试精确匹配
    for col_name in variants:
        if col_name in row.index:
            val = row.get(col_name)
            # 检查是否为 nan 或 None
            if pd.notna(val) and val is not None:
                val_str = str(val).strip()
                # 排除字符串 "nan"（pandas nan 转字符串后的结果）
                if val_str and val_str.lower() != 'nan':
                    return val_str
    
    # 如果精确匹配失败，尝试忽略空格和大小写的匹配
    normalized_row_index = {str(col).strip().lower(): col for col in row.index}
    for variant in variants:
        normalized_variant = variant.strip().lower()
        if normalized_variant in normalized_row_index:
            col_name = normalized_row_index[normalized_variant]
            val = row.get(col_name)
            # 检查是否为 nan 或 None
            if pd.notna(val) and val is not None:
                val_str = str(val).strip()
                # 排除字符串 "nan"（pandas nan 转字符串后的结果）
                if val_str and val_str.lower() != 'nan':
                    return val_str
    return None


def map_excel_row_to_article(row: pd.Series, source_file_name: str, sheet_name: str, row_index: int) -> Optional[dict]:
    """
    将 Excel 行映射为知识条目 DTO（忠实还原原文档模式）
    
    表头：序号 | 现象（问题） | 检查点（原因） | 维修对策（解决办法） | 维修视频（附件）
    """
    # 读取字段（原样保留，不修改）
    # 注意：pandas 读取时，列名可能包含前后空格或特殊字符
    # 支持多种列名格式（带括号和不带括号）
    serial_number = find_column_by_variants(row, ["序号"]) or ""
    
    phenomenon = find_column_by_variants(row, ["现象（问题）", "现象(问题)", "现象 （问题）", "现象 (问题)", "现象", "问题", "故障现象"]) or ""
    
    checkpoints = find_column_by_variants(row, ["检查点（原因）", "检查点(原因)", "检查点 （原因）", "检查点 (原因)", "检查点", "原因"]) or ""
    
    solution = find_column_by_variants(row, ["维修对策（解决办法）", "维修对策(解决办法)", "维修对策 （解决办法）", "维修对策 (解决办法)", "对策", "维修对策", "解决办法", "解决方法"]) or ""
    
    video_reference = find_column_by_variants(row, ["维修视频（附件）", "维修视频(附件)", "维修视频 （附件）", "维修视频 (附件)", "维修视频", "附件", "备注"]) or ""
    
    # 跳过空行和标题行（必须有现象（问题））
    # 如果只有序号或型号等非必需字段，也跳过（避免将标题行或空行录入数据库）
    if not phenomenon:
        return None
    
    # 额外检查：如果现象（问题）的值就是列标题本身（如"现象(问题)"），则跳过
    # 避免将标题行误当作数据行录入
    title_keywords = ["现象", "问题", "检查点", "原因", "维修对策", "解决办法", "维修视频", "附件", "序号", "型号"]
    if phenomenon in title_keywords or any(keyword in phenomenon for keyword in ["现象（问题）", "现象(问题)", "检查点（原因）", "维修对策（解决办法）"]):
        logger.debug(f"跳过标题行或无效行: 现象={phenomenon}")
        return None
    
    # 清理所有字段中的中括号标签（防止Excel单元格中本身就包含这些标签）
    # 同时处理 nan 值：如果字段值是 "nan" 字符串，则清空
    serial_number = clean_bracketed_labels(str(serial_number)) if serial_number and str(serial_number).lower() != 'nan' else ""
    phenomenon = clean_bracketed_labels(phenomenon) if phenomenon and str(phenomenon).lower() != 'nan' else ""
    checkpoints = clean_bracketed_labels(checkpoints) if checkpoints and str(checkpoints).lower() != 'nan' else ""
    solution = clean_bracketed_labels(solution) if solution and str(solution).lower() != 'nan' else ""
    video_reference = clean_bracketed_labels(video_reference) if video_reference and str(video_reference).lower() != 'nan' else ""
    
    # 1) title = 只保留现象（问题），不添加型号信息和中括号
    title = phenomenon.strip()
    # 再次清理，确保没有任何中括号标签
    title = clean_bracketed_labels(title)
    if not title:
        return None
    
    # 2) question_text = 只保留现象（问题）的原始内容，不添加序号、列标题和中括号
    # 序号不录入数据库，只保留现象内容
    question_text = None
    if phenomenon and phenomenon.strip():
        cleaned_phenomenon = clean_bracketed_labels(phenomenon.strip())
        if cleaned_phenomenon:
            question_text = cleaned_phenomenon
    # 最终清理，确保没有任何中括号标签
    if question_text:
        question_text = clean_bracketed_labels(question_text)
    
    # 3) cause_text = 只保留检查点（原因）的原始内容，不添加列标题和中括号
    cause_text = checkpoints.strip() if checkpoints and checkpoints.strip() else None
    # 再次清理，确保没有任何中括号标签
    if cause_text:
        cause_text = clean_bracketed_labels(cause_text)
    
    # 4) solution_text = 保留维修对策（解决办法）和维修视频（附件）的原始内容，不添加列标题和中括号
    # 注意：维修视频（附件）的内容会添加到 solution_text 中，这样即使没有匹配到附件文件，用户也能看到参考信息
    # 同时，如果匹配到附件文件，也会创建附件记录，在附件部分显示
    solution_parts = []
    if solution and solution.strip():
        cleaned_solution = clean_bracketed_labels(solution.strip())
        if cleaned_solution:
            solution_parts.append(cleaned_solution)
    
    # 5) 维修视频（附件）：添加到 solution_text 中（即使没有匹配到附件文件，用户也能看到参考信息）
    if video_reference and video_reference.strip():
        cleaned_video = clean_bracketed_labels(video_reference.strip())
        if cleaned_video:
            solution_parts.append(cleaned_video)
    
    solution_text = "\n\n".join(solution_parts) if solution_parts else None
    # 最终清理，确保没有任何中括号标签
    if solution_text:
        solution_text = clean_bracketed_labels(solution_text)
    
    # 6) scope_json 必须包含：设备系列、来源文件、sheet、行号
    scope_data = {
        "设备系列": "YH400/YH500",
        "来源文件": source_file_name,
        "sheet": sheet_name,
        "行号": row_index
    }
    scope_json = json.dumps(scope_data, ensure_ascii=False)
    
    # 7) tags 至少包含：YH400, YH500, 来源:{文件名}
    tags = ["YH400", "YH500"]
    file_name_without_ext = Path(source_file_name).stem
    if file_name_without_ext:
        tags.append(f"来源:{file_name_without_ext}")
    
    # 如果现象/原因中包含关键字也可追加（简单关键词提取）
    text_content = f"{phenomenon} {checkpoints} {solution}".lower()
    if "yh400" in text_content and "YH400" not in tags:
        # 已经在tags中，跳过
        pass
    if "yh500" in text_content and "YH500" not in tags:
        # 已经在tags中，跳过
        pass
    
    # 提取附件信息（用于后续创建 kb_asset）
    # 支持一个单元格包含多个引用（用换行符、分号等分隔），每个引用都尝试匹配并创建附件记录
    attachment_info_list = []  # 改为列表，支持多个附件
    has_attachment_reference = False  # 标记是否有附件引用（无论是否找到文件）
    if video_reference:
        has_attachment_reference = True
        # 先按换行符分割，然后对每一行再按分号、中文分号等分割
        # 这样可以处理多种格式：换行分隔、分号分隔、或混合
        all_references = []
        for line in video_reference.split('\n'):
            line = line.strip()
            if not line:
                continue
            # 按分号、中文分号分割
            parts = re.split(r'[；;]', line)
            for part in parts:
                part = part.strip()
                if part:
                    all_references.append(part)
        
        # P0: 如果一段里出现多个"参考"，强制二次正则提取并展开
        # 检查是否包含多个"参考"关键字
        reference_count = video_reference.count('参考')
        if reference_count > 1 or len(all_references) == 1:
            # 使用正则表达式提取所有"参考"xxx""格式的引用（支持中文引号 "" 和英文引号 ""）
            # 正则：参考[""]... [""]（支持中文引号和英文引号）
            pattern = r'参考["""""]([^"""""]+)["""""]'
            matches = re.findall(pattern, video_reference)
            if matches:
                # 如果正则提取到更多引用，使用正则提取的结果
                if len(matches) > len(all_references):
                    all_references = [f'参考"{m}"' for m in matches]
                    logger.debug(f"通过正则提取到 {len(all_references)} 个引用（原分割结果: {len(all_references) - len(matches) + len(matches)} 个）")
                elif len(matches) == len(all_references) and reference_count > 1:
                    # 即使数量相同，如果检测到多个"参考"，也使用正则提取的结果（更准确）
                    all_references = [f'参考"{m}"' for m in matches]
                    logger.debug(f"检测到多个'参考'，使用正则提取结果: {len(all_references)} 个引用")
        
        if len(all_references) > 1:
            logger.info(f"检测到 {len(all_references)} 个引用，将尝试匹配所有引用")
        else:
            logger.debug(f"提取到 {len(all_references)} 个引用: {all_references}")
        
        matched_count = 0
        for idx, ref_line in enumerate(all_references, 1):
            extracted_filename = extract_filename_from_reference(ref_line)
            if extracted_filename:
                # 再次清理文件名（确保去除所有引号，包括全角单引号 ''）
                clean_extracted = extracted_filename.strip('"""\'""\'《》【】[]()（）').strip()
                if not clean_extracted:
                    logger.debug(f"引用 {idx}/{len(all_references)}: 清理后文件名为空 (原始: {ref_line[:50]}...)")
                    continue
                
                file_infos = find_attachment_files(clean_extracted)
                if file_infos:
                    # 找到匹配的文件（可能是单个文件或多个文件），全部添加到列表
                    for file_info in file_infos:
                        attachment_info_list.append({
                            "filename": clean_extracted,
                            "file_name": file_info.get("file_name", os.path.basename(file_info["path"])),
                            "url": file_info["url"],
                            "asset_type": file_info["type"],
                            "size": file_info["size"],
                            "relative_path": file_info.get("relative_path"),
                            "source_ref": clean_extracted,  # Excel 引用名（用于溯源）
                            "source_folder": Path(file_info["path"]).parent.name if file_info.get("path") else None
                        })
                    matched_count += len(file_infos)
                    if len(file_infos) == 1:
                        logger.info(f"✓ [{idx}/{len(all_references)}] 找到附件: {clean_extracted} -> {file_infos[0]['url']}")
                    else:
                        logger.info(f"✓ [{idx}/{len(all_references)}] 找到附件（文件夹）: {clean_extracted} -> {len(file_infos)} 个文件")
                else:
                    logger.warning(f"✗ [{idx}/{len(all_references)}] 未找到附件: {clean_extracted} (原始引用: {ref_line[:50]}...)")
            else:
                logger.debug(f"引用 {idx}/{len(all_references)}: 无法提取文件名 (原始: {ref_line[:50]}...)")
        
        if len(all_references) > 1:
            # P1: 匹配统计拆成 "命中引用数 / 总引用数 + 文件总数"
            total_files = len(attachment_info_list)
            logger.info(f"多引用匹配结果: {matched_count}/{len(all_references)} 个引用找到文件，共 {total_files} 个文件")
    
    # _attachment_info 始终是列表（统一处理，避免兼容性问题）
    attachment_info = attachment_info_list if attachment_info_list else None
    
    # 去重：基于 URL 或 relative_path 去重（避免重复创建）
    if attachment_info:
        seen_urls = set()
        unique_attachments = []
        for att in attachment_info:
            url_key = att.get("url") or att.get("relative_path")
            if url_key and url_key not in seen_urls:
                seen_urls.add(url_key)
                unique_attachments.append(att)
        attachment_info = unique_attachments if unique_attachments else None
    
    # 最终检查：确保所有字段都没有中括号标签（最后一次清理）
    final_title = clean_bracketed_labels(title) if title else ""
    final_question_text = clean_bracketed_labels(question_text) if question_text else None
    final_cause_text = clean_bracketed_labels(cause_text) if cause_text else None
    final_solution_text = clean_bracketed_labels(solution_text) if solution_text else None
    
    # 记录清理日志（仅在前几次导入时记录，避免日志过多）
    if row_index <= 3:
        logger.info(f"[行 {row_index}] 清理前 -> 清理后:")
        logger.info(f"  title: {title[:50]}... -> {final_title[:50]}...")
        if question_text:
            logger.info(f"  questionText: {question_text[:50]}... -> {final_question_text[:50] if final_question_text else 'None'}...")
        if cause_text:
            logger.info(f"  causeText: {cause_text[:50]}... -> {final_cause_text[:50] if final_cause_text else 'None'}...")
        if solution_text:
            logger.info(f"  solutionText: {solution_text[:50]}... -> {final_solution_text[:50] if final_solution_text else 'None'}...")
    
    return {
        "title": final_title,
        "questionText": final_question_text,
        "causeText": final_cause_text,
        "solutionText": final_solution_text,
        "scopeJson": scope_json,
        "tags": ", ".join(tags),
        "createdBy": "系统导入",
        "_attachment_info": attachment_info,  # 内部字段，始终是列表或 None，不发送给后端
        "_has_attachment_reference": has_attachment_reference  # 标记是否有附件引用（用于统计）
    }


@app.post("/import/excel", response_model=ExcelImportResponse)
async def import_excel(file: UploadFile = File(...)):
    """
    导入 Excel 文件为知识条目
    
    - 接收 .xlsx 文件
    - 使用 pandas 读取
    - 每行映射为一条知识草稿
    - 调用 .NET 的 /api/ai/kb/articles/batch 创建草稿
    - 如果找到附件文件，创建 kb_asset 记录（方案 B：元数据关联）
    """
    logger.info(f"收到 Excel 导入请求: {file.filename}")
    
    # 验证文件类型
    if not file.filename or not file.filename.endswith('.xlsx'):
        logger.warning(f"文件类型不正确: {file.filename}")
        raise HTTPException(status_code=400, detail="只支持 .xlsx 格式的 Excel 文件")
    
    try:
        # 读取 Excel 文件
        contents = await file.read()
        # 将 bytes 转换为 BytesIO 对象，pandas 才能读取
        excel_file = BytesIO(contents)
        
        # 先尝试从第一行读取表头
        df = pd.read_excel(excel_file, engine='openpyxl', header=0)
        
        # 验证必需字段（支持多种列名）
        # 故障现象列名可能为：故障现象、现象（问题）、现象、问题
        fault_phenomenon_columns = ["故障现象", "现象（问题）", "现象(问题)", "现象 （问题）", "现象 (问题)", "现象", "问题"]
        
        # 改进的列名匹配：支持忽略空格、大小写等
        def normalize_column_name(col_name: str) -> str:
            """标准化列名：去除前后空格，统一大小写"""
            return str(col_name).strip().lower()
        
        # 标准化所有列名和匹配列表
        normalized_columns = {normalize_column_name(col): col for col in df.columns}
        normalized_fault_phenomenon = [normalize_column_name(col) for col in fault_phenomenon_columns]
        
        # 检查是否有匹配的列
        has_fault_phenomenon = any(norm_col in normalized_columns for norm_col in normalized_fault_phenomenon)
        header_row = 0  # 记录 header 行号
        
        logger.info(f"第一行读取到的列名: {list(df.columns)}")
        logger.info(f"标准化后的列名: {list(normalized_columns.keys())}")
        logger.info(f"期望的标准化列名: {normalized_fault_phenomenon}")
        logger.info(f"匹配结果: {has_fault_phenomenon}")
        
        # 如果第一行没有找到必需字段，尝试从第二行读取（跳过标题行）
        if not has_fault_phenomenon:
            logger.info("第一行未找到必需字段，尝试从第二行读取表头（跳过标题行）")
            excel_file.seek(0)  # 重置文件指针
            df = pd.read_excel(excel_file, engine='openpyxl', header=1)
            header_row = 1  # 更新 header 行号
            
            # 重新标准化列名
            normalized_columns = {normalize_column_name(col): col for col in df.columns}
            normalized_fault_phenomenon = [normalize_column_name(col) for col in fault_phenomenon_columns]
            logger.info(f"第二行读取到的列名: {list(df.columns)}")
            logger.info(f"标准化后的列名: {list(normalized_columns.keys())}")
            logger.info(f"期望的标准化列名: {normalized_fault_phenomenon}")
            
            # 再次验证
            has_fault_phenomenon = any(norm_col in normalized_columns for norm_col in normalized_fault_phenomenon)
            logger.info(f"匹配结果: {has_fault_phenomenon}")
            
            # 如果仍然没有匹配，尝试更宽松的匹配（只检查是否包含关键字）
            if not has_fault_phenomenon:
                logger.info("尝试更宽松的匹配：检查列名是否包含'现象'或'问题'关键字")
                for col in df.columns:
                    col_lower = str(col).strip().lower()
                    if '现象' in col_lower or '问题' in col_lower or '故障' in col_lower:
                        logger.info(f"找到可能的匹配列: '{col}' (标准化: '{col_lower}')")
                        has_fault_phenomenon = True
                        break
            
            if not has_fault_phenomenon:
                # 提供更详细的错误信息
                error_msg = (
                    f"缺少必需字段: 请包含以下任一列名 - {', '.join(fault_phenomenon_columns)}\n"
                    f"实际读取到的列名: {', '.join(df.columns)}\n"
                    f"提示: 请确保Excel文件包含'现象（问题）'、'现象(问题)'、'现象 (问题)'或'现象'列"
                )
                logger.error(error_msg)
                raise HTTPException(status_code=400, detail=error_msg)
        
        if df.empty:
            raise HTTPException(status_code=400, detail="Excel 文件为空")
        
        logger.info(f"读取到 {len(df)} 行数据")
        logger.info(f"最终使用的列名: {list(df.columns)}")
        
        # 明确过滤掉"型号"列（如果存在），避免录入数据库
        columns_to_drop = []
        for col in df.columns:
            col_str = str(col).strip()
            # 如果列名包含"型号"，则标记为删除
            if "型号" in col_str:
                columns_to_drop.append(col)
                logger.debug(f"检测到'型号'列，将忽略: {col_str}")
        
        if columns_to_drop:
            df = df.drop(columns=columns_to_drop)
            logger.info(f"已过滤 {len(columns_to_drop)} 个'型号'相关列")
        
        # 获取 sheet 名称（如果可能）
        sheet_name = "Sheet1"  # 默认值
        try:
            # 尝试从 Excel 文件中获取 sheet 名称
            excel_file.seek(0)
            import openpyxl
            wb = openpyxl.load_workbook(excel_file, read_only=True)
            if wb.sheetnames:
                sheet_name = wb.sheetnames[0]  # 使用第一个 sheet 的名称
            excel_file.seek(0)  # 重置文件指针
        except Exception as e:
            logger.warning(f"无法获取 sheet 名称，使用默认值: {str(e)}")
        
        # 确定 header 行号（用于计算 Excel 行号）
        header_row = 0 if has_fault_phenomenon else 1
        
        # 映射每行为知识条目 DTO
        articles = []  # 发送给后端的文章（不包含内部字段）
        articles_with_attachments = []  # 包含附件信息的完整文章数据
        failures = []
        skipped_rows = 0
        attachment_match_count = 0  # 附件匹配统计
        attachment_not_found_count = 0  # 附件未找到统计
        
        for idx, row in df.iterrows():
            try:
                # Excel 行号 = pandas index + header行数 + 1（Excel从1开始计数）
                # 如果 header=0，则 Excel行号 = idx + 2（第1行是表头，第2行开始是数据）
                # 如果 header=1，则 Excel行号 = idx + 3（第1行是标题，第2行是表头，第3行开始是数据）
                excel_row_number = int(idx) + header_row + 2
                
                # 调试：打印第一行的数据
                if idx == 0:
                    logger.info(f"第一行数据预览: {dict(row)}")
                
                article = map_excel_row_to_article(row, file.filename, sheet_name, excel_row_number)
                if article:
                    # 保存包含附件信息的原始 article
                    articles_with_attachments.append(article)
                    # 移除内部字段（不发送给后端）
                    article_for_api = {k: v for k, v in article.items() if not k.startswith("_")}
                    articles.append(article_for_api)
                else:
                    skipped_rows += 1
                    if idx < 3:  # 只记录前3行的跳过原因
                        logger.info(f"第 {excel_row_number} 行被跳过（空行或缺少必需字段）")
                # 如果返回 None，说明是空行，跳过
            except Exception as e:
                excel_row_number = int(idx) + header_row + 2
                logger.error(f"处理第 {excel_row_number} 行时出错: {str(e)}")
                failures.append({
                    "row_index": excel_row_number,
                    "reason": str(e)
                })
        
        # 统计附件匹配情况（支持多个附件）
        attachment_match_count = 0
        for a in articles_with_attachments:
            att_info = a.get("_attachment_info")
            if att_info:
                # _attachment_info 现在始终是列表
                if isinstance(att_info, list):
                    attachment_match_count += len(att_info)
                else:
                    # 兼容旧数据
                    attachment_match_count += 1
        
        attachment_total = sum(1 for a in articles_with_attachments if a.get("_has_attachment_reference", False))
        attachment_not_found_count = attachment_total - sum(1 for a in articles_with_attachments if a.get("_attachment_info"))
        
        logger.info(f"成功映射 {len(articles)} 条，跳过 {skipped_rows} 行，失败 {len(failures)} 行")
        if attachment_total > 0:
            match_rate = (attachment_match_count / attachment_total * 100) if attachment_total > 0 else 0
            logger.info(f"📎 附件匹配统计: 找到 {attachment_match_count}/{attachment_total} 个 ({match_rate:.1f}%)，未找到 {attachment_not_found_count} 个")
        
        if not articles:
            error_msg = f"没有有效的数据行。总行数: {len(df)}, 跳过: {skipped_rows}, 失败: {len(failures)}"
            if len(df) > 0:
                error_msg += f"\n列名: {list(df.columns)}"
                error_msg += f"\n第一行数据: {dict(df.iloc[0]) if len(df) > 0 else 'N/A'}"
            raise HTTPException(status_code=400, detail=error_msg)
        
        # 调用 .NET 后端批量创建接口
        logger.info(f"准备调用 .NET 后端: {DOTNET_BASE_URL}/api/ai/kb/articles/batch, 文章数量: {len(articles)}")
        
        async with httpx.AsyncClient() as client:
            try:
                response = await client.post(
                    f"{DOTNET_BASE_URL}/api/ai/kb/articles/batch",
                    json={"articles": articles},
                    headers={
                        "Content-Type": "application/json",
                        "X-Tenant-Id": DEFAULT_TENANT,
                        "X-Internal-Token": INTERNAL_TOKEN
                    },
                    timeout=30.0
                )
                
                logger.info(f".NET 后端响应状态码: {response.status_code}")
                
                if response.status_code != 200:
                    error_text = response.text
                    logger.error(f".NET 后端返回错误: {error_text}")
                    raise HTTPException(
                        status_code=response.status_code,
                        detail=f".NET 后端返回错误: {error_text}"
                    )
            except httpx.ConnectError as e:
                logger.error(f"无法连接到 .NET 后端 {DOTNET_BASE_URL}: {str(e)}")
                raise HTTPException(
                    status_code=503,
                    detail=f"无法连接到 .NET 后端服务，请确保服务已启动: {DOTNET_BASE_URL}"
                )
            except httpx.TimeoutException as e:
                logger.error(f"调用 .NET 后端超时: {str(e)}")
                raise HTTPException(
                    status_code=504,
                    detail="调用 .NET 后端超时，请稍后重试"
                )
            
            result = response.json()
            
            # 处理返回结果
            success_count = result.get("successCount", 0)
            failure_count = result.get("failureCount", 0)
            article_ids = []
            article_id_to_attachment = {}  # 映射 article_id -> attachment_info
            
            # 收集成功创建的 article IDs 和附件信息
            # 支持一个文章有多个附件（_attachment_info 可能是单个对象或列表）
            article_id_to_attachments = {}  # 改为映射到附件列表
            for i, item in enumerate(result.get("results", [])):
                if item.get("success") and item.get("articleId"):
                    article_id = item["articleId"]
                    article_ids.append(article_id)
                    
                    # 如果对应的 article 有附件信息，保存映射
                    if i < len(articles_with_attachments):
                        att_info = articles_with_attachments[i].get("_attachment_info")
                        if att_info:
                            # _attachment_info 现在始终是列表（统一处理）
                            if isinstance(att_info, list):
                                article_id_to_attachments[article_id] = att_info
                            else:
                                # 兼容旧数据（理论上不应该出现，但保险起见）
                                article_id_to_attachments[article_id] = [att_info]
                elif not item.get("success"):
                    # 记录后端返回的失败信息
                    failures.append({
                        "row_index": item.get("index", -1) + (header_row + 2),  # 转换为 Excel 行号
                        "reason": item.get("error", "未知错误")
                    })
            
            # 批量创建附件（方案 B：元数据关联）
            if article_id_to_attachments:
                total_attachments = sum(len(atts) for atts in article_id_to_attachments.values())
                logger.info(f"准备创建 {total_attachments} 个附件记录（涉及 {len(article_id_to_attachments)} 篇文章）")
                
                # 记录每个文章的附件数量
                for article_id, att_info_list in article_id_to_attachments.items():
                    logger.info(f"  文章 ID {article_id}: {len(att_info_list)} 个附件")
                
                assets_to_create = []
                
                for article_id, att_info_list in article_id_to_attachments.items():
                    # 为每个附件创建记录
                    for att_info in att_info_list:
                        assets_to_create.append({
                            "articleId": article_id,
                            "assetType": att_info["asset_type"],
                            "fileName": att_info["file_name"],
                            "url": att_info["url"],
                            "size": att_info["size"],
                            "duration": None  # 视频时长暂不支持自动获取
                        })
                
                if assets_to_create:
                    try:
                        asset_response = await client.post(
                            f"{DOTNET_BASE_URL}/api/ai/kb/articles/assets/batch",
                            json={"assets": assets_to_create},
                            headers={
                                "Content-Type": "application/json",
                                "X-Tenant-Id": DEFAULT_TENANT,
                                "X-Internal-Token": INTERNAL_TOKEN
                            },
                            timeout=30.0
                        )
                        
                        if asset_response.status_code == 200:
                            asset_result = asset_response.json()
                            success_count = asset_result.get('successCount', 0)
                            failure_count = asset_result.get('failureCount', 0)
                            logger.info(f"✅ 附件创建结果: 成功 {success_count} 个，失败 {failure_count} 个")
                            
                            # 记录成功的附件详情（包含数据库ID）
                            if success_count > 0:
                                results = asset_result.get('results', [])
                                created_assets = []  # 记录成功创建的附件信息
                                for result_item in results:
                                    if result_item.get('success'):
                                        index = result_item.get('index', -1)
                                        asset_id = result_item.get('assetId')
                                        if index < len(assets_to_create):
                                            success_asset = assets_to_create[index]
                                            created_assets.append({
                                                'articleId': success_asset.get('articleId'),
                                                'fileName': success_asset.get('fileName'),
                                                'assetId': asset_id
                                            })
                                            logger.info(f"  ✓ 附件 [{index}] 已写入数据库: ArticleId={success_asset.get('articleId')}, FileName={success_asset.get('fileName')}, AssetId={asset_id}")
                                
                                # 验证：尝试查询刚创建的附件（可选，用于确认）
                                if created_assets:
                                    logger.info(f"📋 附件记录已成功写入数据库，共 {len(created_assets)} 条记录")
                            
                            # 记录失败的附件详情
                            if failure_count > 0:
                                results = asset_result.get('results', [])
                                for result_item in results:
                                    if not result_item.get('success'):
                                        index = result_item.get('index', -1)
                                        error = result_item.get('error', '未知错误')
                                        if index < len(assets_to_create):
                                            failed_asset = assets_to_create[index]
                                            logger.warning(f"  ✗ 附件创建失败 [{index}]: ArticleId={failed_asset.get('articleId')}, FileName={failed_asset.get('fileName')}, Error={error}")
                        else:
                            logger.error(f"❌ 创建附件API调用失败: {asset_response.status_code} - {asset_response.text}")
                            logger.error(f"   响应内容: {asset_response.text[:500]}")
                    except Exception as e:
                        logger.error(f"创建附件时出错: {str(e)}")
                        # 附件创建失败不影响主流程
            
            # 合并前端解析失败和后端创建失败
            total_failures = len(failures) + failure_count
            
            return ExcelImportResponse(
                total_rows=len(df),
                success_count=success_count,
                failure_count=total_failures,
                article_ids=article_ids,
                failures=failures
            )
    
    except pd.errors.EmptyDataError:
        logger.error("Excel 文件为空或格式不正确")
        raise HTTPException(status_code=400, detail="Excel 文件为空或格式不正确")
    except httpx.HTTPError as e:
        logger.error(f"调用 .NET 后端失败: {str(e)}")
        logger.error(traceback.format_exc())
        raise HTTPException(
            status_code=500,
            detail=f"调用 .NET 后端失败: {str(e)}"
        )
    except Exception as e:
        logger.error(f"导入失败: {str(e)}")
        logger.error(traceback.format_exc())
        raise HTTPException(
            status_code=500,
            detail=f"导入失败: {str(e)}"
        )


@app.get("/health")
async def health():
    """健康检查"""
    return {"status": "ok", "service": "ai-hub-ai"}


if __name__ == "__main__":
    import uvicorn
    import os
    port = int(os.getenv("PORT", "8000"))
    uvicorn.run(app, host="0.0.0.0", port=port)
